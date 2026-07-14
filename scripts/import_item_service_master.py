#!/usr/bin/env python3
"""Import a client item/service master XLSX into a local SQLite lookup DB."""

from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "openpyxl is required to import XLSX files. Install it or export the file as CSV first."
    ) from exc


EXPECTED_HEADERS = [
    "ItemCode",
    "ItemDesc",
    "BGCode",
    "BGDesc",
    "BSGCode",
    "BSGDesc",
    "ActiveDateTo",
    "Category",
]


def normalize_text(value: object) -> str:
    text = str(value or "").upper()
    text = text.replace("&", " AND ")
    text = text.replace("X-RAY", "XRAY")
    text = text.replace("X RAY", "XRAY")
    text = text.replace("2 D ECHO", "2DECHO")
    text = text.replace("2D ECHO", "2DECHO")
    text = re.sub(r"\b(\d+)\s*K\b", lambda match: str(int(match.group(1)) * 1000), text)
    text = re.sub(r"([A-Z])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([A-Z])", r"\1 \2", text)
    text = re.sub(r"\bHBA\s+1\s+C\b", "HBA1C", text)
    text = re.sub(r"\bA\s+1\s+C\b", "A1C", text)
    text = re.sub(r"\b2\s+DECHO\b", "2DECHO", text)
    chars = []
    for char in text:
        chars.append(char if char.isalnum() else " ")
    return " ".join("".join(chars).split())


def normalize_date(value: object) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[1]
    default_db = repo_root / "server" / "storage" / "item_service_master.sqlite"

    parser = argparse.ArgumentParser(
        description="Build a local SQLite lookup database from item_service_master_active XLSX."
    )
    parser.add_argument("--source", required=True, help="Path to the client XLSX file.")
    parser.add_argument("--db", default=str(default_db), help="Output SQLite DB path.")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to the first sheet.")
    parser.add_argument("--limit", type=int, default=None, help="Import only N data rows.")
    return parser.parse_args()


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode = OFF;
        PRAGMA synchronous = OFF;

        CREATE TABLE item_service_master (
          id INTEGER PRIMARY KEY,
          item_code TEXT NOT NULL,
          item_desc TEXT NOT NULL,
          bg_code TEXT,
          bg_desc TEXT,
          bsg_code TEXT,
          bsg_desc TEXT,
          active_date_to TEXT,
          category TEXT,
          normalized_desc TEXT NOT NULL
        );

        CREATE VIRTUAL TABLE item_service_master_fts USING fts5(
          item_code UNINDEXED,
          normalized_desc,
          item_desc UNINDEXED,
          bg_code UNINDEXED,
          bg_desc UNINDEXED,
          bsg_code UNINDEXED,
          bsg_desc UNINDEXED,
          active_date_to UNINDEXED,
          category UNINDEXED
        );

        CREATE TABLE catalog_meta (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        );
        """
    )


def insert_batch(conn: sqlite3.Connection, rows: list[tuple]) -> None:
    conn.executemany(
        """
        INSERT INTO item_service_master (
          item_code,
          item_desc,
          bg_code,
          bg_desc,
          bsg_code,
          bsg_desc,
          active_date_to,
          category,
          normalized_desc
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    conn.executemany(
        """
        INSERT INTO item_service_master_fts (
          item_code,
          normalized_desc,
          item_desc,
          bg_code,
          bg_desc,
          bsg_code,
          bsg_desc,
          active_date_to,
          category
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [(row[0], row[8], row[1], row[2], row[3], row[4], row[5], row[6], row[7]) for row in rows],
    )


def import_workbook(source: Path, db_path: Path, sheet_name: str | None, limit: int | None) -> int:
    if not source.exists():
        raise FileNotFoundError(source)

    db_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_db_name = tempfile.mkstemp(
        prefix=f"{db_path.stem}.",
        suffix=".tmp",
        dir=str(db_path.parent),
    )
    os.close(fd)
    temp_db = Path(temp_db_name)

    count = 0
    try:
        conn = sqlite3.connect(temp_db)
        create_schema(conn)

        workbook = load_workbook(source, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
            raise ValueError(f"Unexpected headers: {headers}")

        batch: list[tuple] = []
        for row in rows:
            if limit is not None and count >= limit:
                break

            item_code = row[0]
            item_desc = row[1]
            if item_code is None or item_desc is None or str(item_desc).strip() == "":
                continue

            normalized_desc = normalize_text(item_desc)
            batch.append(
                (
                    str(item_code).strip(),
                    str(item_desc).strip(),
                    str(row[2]).strip() if row[2] is not None else None,
                    str(row[3]).strip() if row[3] is not None else None,
                    str(row[4]).strip() if row[4] is not None else None,
                    str(row[5]).strip() if row[5] is not None else None,
                    normalize_date(row[6]),
                    str(row[7]).strip() if row[7] is not None else None,
                    normalized_desc,
                )
            )
            count += 1

            if len(batch) >= 5000:
                insert_batch(conn, batch)
                conn.commit()
                batch = []

        if batch:
            insert_batch(conn, batch)
            conn.commit()

        conn.executescript(
            """
            CREATE INDEX idx_item_service_master_bg_desc ON item_service_master(bg_desc);
            CREATE INDEX idx_item_service_master_bsg_desc ON item_service_master(bsg_desc);
            CREATE INDEX idx_item_service_master_category ON item_service_master(category);
            CREATE INDEX idx_item_service_master_item_code ON item_service_master(item_code);
            CREATE INDEX idx_item_service_master_normalized_desc ON item_service_master(normalized_desc);
            """
        )
        conn.executemany(
            "INSERT INTO catalog_meta(key, value) VALUES (?, ?)",
            [
                ("source_path", str(source)),
                ("sheet_name", worksheet.title),
                ("imported_at", datetime.now(timezone.utc).replace(microsecond=0).isoformat()),
                ("row_count", str(count)),
            ],
        )
        conn.commit()
        conn.close()
        os.replace(temp_db, db_path)
        return count
    except Exception:
        try:
            temp_db.unlink(missing_ok=True)
        finally:
            raise


def main() -> int:
    args = parse_args()
    source = Path(args.source).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()
    count = import_workbook(source, db_path, args.sheet, args.limit)
    print(f"Imported {count} rows into {db_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
